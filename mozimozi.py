import numpy as np
import pandas as pd
from rapidfuzz.process import cdist
from tqdm import tqdm
import xlsxwriter

# 疑似データを生成 (コメントアウト)
# from faker import Faker
# fake = Faker('ja-JP')
# N_target = 5000
# N_db = 1000
# target = pd.Series([fake.company() for _ in range(N_target)])
# db = pd.Series([fake.company() for _ in range(N_db)]).sort_values()

# エクセルからターゲットとデータベースを読み込む
target_path = 'target.xlsx'  # ターゲットのExcelファイルのパス
db_path = 'db.xlsx'  # データベースのExcelファイルのパス
result_path = 'result.xlsx'  # 結果のExcelファイルのパス
target = pd.read_excel(target_path, sheet_name='Sheet1')['company_name']  # シート名や列名は適宜修正
target_name = pd.DataFrame(
    pd.read_excel(target_path, sheet_name='Sheet1')['name'], columns=['name'])
db = pd.read_excel(db_path, sheet_name='Sheet1')['company_name'].sort_values()  # シート名や列名は適宜修正


# 先頭5件を出力
print('データの読み込み完了、先頭5件を出力します。')
print(target[:5].to_list())
print(db[:5].to_list())

# スコアを計算
print('類似度を計算中...')
score = cdist(target, db)
df = pd.DataFrame(score, index=target, columns=db)

# 各行の最大値、2番目、3番目の値を取得
df_score = pd.DataFrame()
df_score[['max_value', 'second_value', 'third_value']] = df.apply(
    lambda row: row.nlargest(3).values,
    axis=1,
    result_type='expand'  # 複数列に展開
)

# 各行の最大値、2番目、3番目のインデックスを取得
df_score[['max_index', 'second_index', 'third_index']] = df.apply(
    lambda row: row.nlargest(3).index,
    axis=1,
    result_type='expand'  # 複数列に展開
)

# 結果を結合
df = pd.concat([df_score, df], axis=1)
df = df.reset_index(drop=False)
df = pd.concat([target_name, df], axis=1)
print('類似会社名を取得完了、先頭5件を出力します。')
print(df.head())


# 結果をエクセルに出力
print('結果をエクセルに出力します。これには時間がかかることがあります。')
df.to_excel(result_path, sheet_name='Sheet1')
'''
# エクセルファイルへの書き出し (プログレスバー付き)
with pd.ExcelWriter('result.xlsx', engine='xlsxwriter') as writer:
    # データを書き出すためのワークシートを追加
    df.to_excel(writer, sheet_name='new_sheet_name', index=True, startrow=0, header=False)
    # ワークシートを取得
    worksheet = writer.sheets['new_sheet_name']
    # 列名を書き出し
    for col_num, col_name in enumerate(df.columns):
        worksheet.write(0, col_num, col_name)
    # 行データの書き出し
    for row_num, (index, row) in enumerate(tqdm(df.iterrows(), total=len(df), desc="Writing to Excel")):
        for col_num, value in enumerate(row):
            worksheet.write(row_num + 1, col_num, value)
'''

print('処理が完了しました。')

# excelファイルの編集
import openpyxl
from openpyxl.styles import PatternFill

print('エクセルファイルの編集を開始します。')
# 既存のExcelファイルを開く
workbook = openpyxl.load_workbook(result_path)
sheet = workbook['Sheet1']  # シート名を指定

# 黄色のハッチパターンを定義
hatch_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='lightUp')

# ハッチを追加したい列のリスト（例: A列、C列、E列）
columns_to_hatch = ['C', 'G']  # とびとびの列名を指定

# 指定した列にハッチを追加
for col in columns_to_hatch:
    for row in range(1, sheet.max_row + 1):  # 1行目から最終行まで
        cell = sheet[f'{col}{row}']  # セルを取得
        cell.fill = hatch_fill  # ハッチパターンを適用

print('ハッチパターンを追加しました。')

# 列の幅を調整（A列: 20, B列: 30, C列: 15）
column_widths = {
    'B': 20,
    'C': 30,
    'G': 30
}

# 指定した列の幅を設定
for col, width in column_widths.items():
    sheet.column_dimensions[col].width = width

print('列の幅を調整しました。')

# 変更を保存
workbook.save(file_path)
workbook.close()
